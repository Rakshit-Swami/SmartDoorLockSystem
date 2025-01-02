from tkinter import * # type: ignore
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl , xlrd
from openpyxl import Workbook
import pathlib

 

root=Tk()
root.title("Smart Door System")
root.geometry("986x640+200+100")
root.resizable(False,False)

####################Registration NO.#######################
#it is created to automatic enter registration no. ,It will update registration number on each save

def registration_no():
    file=openpyxl.load_workbook('Database/person list.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value

    try:
        Registration.set(max_row_value+1)

    except:
        Registration.set("1")   #if we are using for first time, then it will take 1 as registration number 


#######################Automatic excel File Creation(addition of data column),if it is not available######################3 
file=pathlib.Path('Database/person list.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No."
    sheet['B1']="Name"
    sheet['C1']="Rollno"


    file.save('Database/person list.xlsx')
########################ShowImage#########################
def showimage():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),
                                        title="Select image file",filetype=(("JPG File","*.jpg"),
                                                                             ("PNG File","*.png"),
                                                                             ("All files","*.txt")))
    img = (Image.open(filename))
    resized_image= img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

###########################Save#####################
def Save():
    R1=Registration.get()
    N1=Name.get()
    


    D2=Rollno.get()

   

    if N1=="" or D2=="":
        messagebox.showerror("error","Few Data is missing!")

    else:
        file=openpyxl.load_workbook('Database/person list.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=N1)
        sheet.cell(column=3,row=sheet.max_row,value=D2)
       
        
        file.save(r'Database/person list.xlsx')

        try:
            img.save("Access/"+str(N1+D2)+".jpg")
        except:
            messagebox.showinfo("info","Profile Picture is not available!!!!")

        messagebox.showinfo("info","Sucessfully data entered!!!")

        Clear()      #clear entry box and image section

        registration_no()  #it will recheck registration no. and re-issue new no.

##################Clear#######################
#to clear all the entry feild , so that we can implement new
def Clear():
    global img
    Name.set('')
    Rollno.set('')
   

    registration_no()  #after clear it will check for registration number again


    img1=PhotoImage(file='Images/logo.png')
    lbl.config(image=img1)
    lbl.image=img1

    img=""     #image is set to be blank , so that it will not use same image for next entry



#icon
image_icon=PhotoImage(file="Images/logo.png")
root.iconphoto(False,image_icon)

back_image=PhotoImage(file="Images/back.png")
Label(root,image=back_image).pack()

#image
f=Frame(root,bd=1,bg="black",width=170,height=150,relief=GROOVE)
f.place(x=450,y=150)


Name=StringVar()
name_entry = Entry(root,textvariable=Name,width=20,font="arial 9",bd=0,bg="#dbe0e3",justify=CENTER)
name_entry.place(x=470,y=330)

Rollno=StringVar()
rollno_entry = Entry(root,textvariable=Rollno,width=20,font="arial 9",bd=0,bg="#dbe0e3",justify=CENTER)
rollno_entry.place(x=470,y=358)

Registration=IntVar()
Date = StringVar()

reg_entry = Entry(root,textvariable=Registration,width=15,font="arial 10")
reg_entry.place(x=160,y=150)

registration_no()

img=PhotoImage(file="Images/logo.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)

Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showimage).place(x=670,y=340)
saveButton=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightgreen",command=Save)
saveButton.place(x=670,y=410) 

root.mainloop()
